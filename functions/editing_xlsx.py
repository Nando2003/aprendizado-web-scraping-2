from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def editing_xlsx(excel_path: str, data: list, linha: int = 1, coluna: str = 'A') -> bool:
    """
    Preenche uma coluna de um arquivo Excel com dados fornecidos.
    
    :param excel_path: Caminho do arquivo Excel.
    :param data: Lista de dados para preencher. Pode ser [(data, color), (data, color)...], [data, data...] ou [(data, color), data...]
    :param linha: Linha inicial para preenchimento (1-indexado).
    :param coluna: Coluna para preenchimento.
    :return: True se a operação foi bem-sucedida, False caso contrário.
    """
    try:
        load_wb = load_workbook(excel_path)
        load_ws = load_wb.active
        
        coluna = coluna.upper()
        linha = linha - 1
        for i, item in enumerate(data, start=linha):
            cell = load_ws[f'{coluna}{i+1}']
            if isinstance(item, tuple):
                value, color = item
                cell.value = value
                if color:
                    cell.fill = PatternFill(
                        start_color=color,
                        end_color=color,
                        fill_type='solid'
                    )
            else:
                cell.value = item
        
        load_wb.save(excel_path)
        return True
        
    except Exception as e:
        return False
